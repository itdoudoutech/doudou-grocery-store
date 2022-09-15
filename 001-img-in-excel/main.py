# -*- coding: utf-8 -*-
import os
from PIL import Image
import openpyxl
import openpyxl.styles
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


def rgb2hex(rgb):
    print(rgb)
    rgb = rgb.split(',')
    color = ''
    for i in rgb:
        num = int(i)
        color += str(hex(num))[-2:].replace('x', '0').upper()
    return color


def img2excel(img_path, excel_path):
    img_src = Image.open(img_path)
    # 图片宽高
    img_width = img_src.size[0]
    img_height = img_src.size[1]

    str_strlist = img_src.load()
    wb = openpyxl.Workbook()
    wb.save(excel_path)
    wb = openpyxl.load_workbook(excel_path)
    cell_width, cell_height = 1.0, 1.0

    sheet = wb["Sheet"]
    for w in range(img_width):
        for h in range(img_height):
            data = str_strlist[w, h]
            color = str(data).replace("(", "").replace(")", "")
            color = rgb2hex(color)
            # 设置填充颜色为 color
            fille = PatternFill("solid", fgColor=color)
            sheet.cell(h + 1, w + 1).fill = fille
    for i in range(1, sheet.max_row + 1):
        sheet.row_dimensions[i].height = cell_height
    for i in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(i)].width = cell_width
    wb.save(excel_path)
    img_src.close()


if __name__ == '__main__':
    path = os.path.split(os.path.realpath(__file__))[0]
    img_path = F'{path}/fire.jpg'
    excel_path = F'{path}/fire.xlsx'
    img2excel(img_path, excel_path)
